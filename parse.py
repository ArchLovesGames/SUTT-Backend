#import json and openpyxl
import json
import openpyxl

#load the excel sheet
wb_obj = openpyxl.load_workbook("Mess Menu Sample.xlsx")
#activate the excel sheet for reading the data
sheet_obj = wb_obj.active

#create a list with the days and meals to add to the 'OMIT' section
days = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]
headers = ["BREAKFAST", "LUNCH", "DINNER"]
#initialise rows and columns and menu_list, the list which will be added to the json file
rows = sheet_obj.max_row
cols = sheet_obj.max_column
menu_list = []

#object for the menu
class Menu:
    def __init__(self, date, bitems, litems, ditems):
        self.DATE = date
        self.BREAKFAST = bitems
        self.LUNCH = litems
        self.DINNER = ditems

#iteration to start gathering the data
i = 1
#initialise the lists for the individual meal items for given date
while i < cols:
    blist, llist, dlist = [], [], []
    r = 2
    d = sheet_obj.cell(r, i).value
    r += 1

#start loop for given date 
    while r < rows + 1:
#append breakfast items to blist
        while r < rows and sheet_obj.cell(r, i).value not in days:
            cell_value = sheet_obj.cell(r, i).value
            if cell_value in headers or cell_value == None or cell_value == "**************" or cell_value == "***************":
                r += 1
                continue
            blist.append(cell_value)
            r += 1

        r += 1
#appending lunch items
        while r < rows and sheet_obj.cell(r, i).value not in days:
            cell_value = sheet_obj.cell(r, i).value
            if cell_value in headers or cell_value == None or cell_value == "**************" or cell_value == "***************":
                r += 1
                continue
            llist.append(cell_value)
            r += 1
#appending dinner items
        r += 1
        while r < rows:
            cell_value = sheet_obj.cell(r, i).value
            if cell_value in headers or cell_value == None or cell_value == "**************" or cell_value == "***************":
                r += 1
                continue
            dlist.append(cell_value)
            r += 1
#appending all the sorted items to the menu list
        menu_list.append(Menu(d, blist, llist, dlist))
        r += 1
#move to the next date
    i += 1

#initialise json list and append the menu items in menu_list, after sorting them in the MENU object accordingly, to the json list
menu_json = []
for menu in menu_list:
    menu_json.append({
        menu.DATE.strftime("%d-%m-%Y"): {
            "Breakfast": menu.BREAKFAST,
            "Lunch": menu.LUNCH,
            "Dinner": menu.DINNER
        }
    })
  
#send appended list to menu.json
with open("menu.json", "w") as file:
    json.dump(menu_json, file)

#be happy as you realise you have 2 tut tests the next day



    


    
  
  
  
  
